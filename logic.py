"""
Модуль бизнес-логики для системы автоматизации отчетов
Версия с поддержкой типов отчётов (месячный, квартальный, годовой)
"""

import os
import re
import openpyxl
from database import save_report_to_db, get_all_reports, get_report_by_id, delete_report
from export_excel import create_excel_report


def sanitize_filename(filename):
    """
    Очистить имя файла от символов, запрещённых в Windows
    Заменяет : / \\ * ? " < > | на _
    """
    forbidden_chars = r'[:\\/*?\"<>|]'
    cleaned = re.sub(forbidden_chars, '_', filename)
    # Убираем множественные подчёркивания
    cleaned = re.sub(r'_{2,}', '_', cleaned)
    # Убираем подчёркивания в начале и конце
    cleaned = cleaned.strip('_')
    return cleaned


def get_report_type_by_month(month):
    """
    Определить тип отчёта по месяцу
    1 - месячный
    2 - квартальный (март, июнь, сентябрь, декабрь)
    3 - годовой (январь)
    """
    month_lower = month.lower()

    # Годовой отчёт - январь
    if month_lower in ['январь', 'january']:
        return 3

    # Квартальные отчёты
    quarterly_months = ['март', 'march', 'июнь', 'june', 'сентябрь', 'september', 'декабрь', 'december']
    if month_lower in quarterly_months:
        return 2

    # Месячный отчёт - все остальные
    return 1


def find_form_file(form_name, month):
    """
    Найти файл формы с учётом типа отчёта
    Ищет файлы в порядке:
    1. [Должность]_[тип].xlsx (например, Главный_инженер_2.xlsx)
    2. [Должность].xlsx (если файла с типом нет)

    Возвращает: (полный_путь, название_файла) или (None, None) если не найдено
    """
    forms_dir = "формы"

    if not os.path.exists(forms_dir):
        return None, None

    # Определяем тип отчёта
    report_type = get_report_type_by_month(month)

    # Пробуем файл с типом
    typed_filename = f"{form_name}_{report_type}.xlsx"
    typed_path = os.path.join(forms_dir, typed_filename)

    if os.path.exists(typed_path):
        return typed_path, typed_filename

    # Fallback - файл без типа
    fallback_filename = f"{form_name}.xlsx"
    fallback_path = os.path.join(forms_dir, fallback_filename)

    if os.path.exists(fallback_path):
        return fallback_path, fallback_filename

    return None, None


class ReportLogic:
    """Класс с бизнес-логикой приложения"""

    def __init__(self):
        self.current_report_data = {}
        self.questions_list = []
        self.answers_list = []
        self.current_question_index = 0
        self.questions_per_page = 5

    def load_forms_list(self):
        """Загрузка списка форм из папки 'формы/'"""
        forms_dir = "формы"
        if not os.path.exists(forms_dir):
            return []
        files = os.listdir(forms_dir)
        excel_files = [f for f in files if f.endswith(('.xlsx', '.xls'))]
        # Убираем цифры из имён (Должность_1.xlsx -> Должность)
        forms_set = set()
        for f in excel_files:
            name = os.path.splitext(f)[0]
            # Убираем _1, _2, _3 если есть
            base_name = re.sub(r'_[123]$', '', name)
            forms_set.add(base_name)
        return sorted(list(forms_set))

    def load_questions_from_excel(self, file_path):
        """
        Загрузка вопросов из Excel файла
        Возвращает True при успехе, False при ошибке
        """
        try:
            # Проверка существования файла
            if not os.path.exists(file_path):
                print(f"Ошибка: файл не найден: {file_path}")
                return False

            # Проверка расширения файла
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                print(f"Ошибка: неподдерживаемый формат файла. Ожидается .xlsx или .xls")
                return False

            # Загрузка Excel файла
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"Ошибка: файл повреждён или имеет неверный формат Excel")
                return False
            except PermissionError:
                print(f"Ошибка: нет доступа к файлу. Возможно, он открыт в другой программе")
                return False

            ws = wb.active

            # Проверка что лист не пустой
            if ws.max_row < 2:
                print(f"Ошибка: файл не содержит данных (нужно минимум 2 строки)")
                wb.close()
                return False

            self.questions_list = []
            rows_processed = 0
            rows_skipped = 0

            # Читаем данные начиная со 2-й строки (1-я строка - заголовки)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Пропускаем полностью пустые строки
                if not any(row):
                    continue

                # Проверяем наличие вопроса в первой колонке
                if not row[0]:
                    rows_skipped += 1
                    print(f"Предупреждение: строка {row_num} пропущена - отсутствует текст вопроса")
                    continue

                # Добавляем вопрос
                self.questions_list.append({
                    'question': str(row[0]).strip() if row[0] else "",
                    'gost': str(row[1]).strip() if len(row) > 1 and row[1] else "",
                    'quality': str(row[2]).strip() if len(row) > 2 and row[2] else "",
                    'documents': str(row[3]).strip() if len(row) > 3 and row[3] else ""
                })
                rows_processed += 1

            wb.close()

            # Проверка что загружен хотя бы один вопрос
            if len(self.questions_list) == 0:
                print(f"Ошибка: в файле не найдено ни одного валидного вопроса")
                return False

            print(f"Успешно загружено {rows_processed} вопросов")
            if rows_skipped > 0:
                print(f"Пропущено {rows_skipped} строк без вопросов")

            return True

        except Exception as e:
            print(f"Неожиданная ошибка при загрузке Excel: {type(e).__name__}: {e}")
            return False

    def init_report(self, form_name, month, year, report_date):
        """Инициализация нового отчета"""
        self.current_report_data = {
            'form_name': form_name,
            'month': month,
            'year': year,
            'report_date': report_date
        }

        self.answers_list = [{
            'question_text': q['question'],
            'answer_yes_no': '',
            'comment': '',
            'gost_text': q['gost'],
            'quality_text': q['quality'],
            'documents_text': q['documents']
        } for q in self.questions_list]

        self.current_question_index = 0
        return True

    def get_current_block_questions(self):
        """Получить вопросы текущего блока"""
        start = self.current_question_index
        end = min(start + self.questions_per_page, len(self.questions_list))
        return start, end

    def save_answer(self, question_index, answer_yes_no, comment):
        """Сохранить ответ на вопрос"""
        if 0 <= question_index < len(self.answers_list):
            self.answers_list[question_index]['answer_yes_no'] = answer_yes_no
            self.answers_list[question_index]['comment'] = comment
            return True
        return False

    def check_all_answered(self):
        """Проверить что все вопросы отвечены"""
        for i, answer in enumerate(self.answers_list):
            if not answer['answer_yes_no']:
                return False, i + 1
        return True, None

    def next_block(self):
        """Переход к следующему блоку"""
        next_start = self.current_question_index + self.questions_per_page
        if next_start >= len(self.questions_list):
            return False
        self.current_question_index = next_start
        return True

    def prev_block(self):
        """Возврат к предыдущему блоку"""
        prev_start = self.current_question_index - self.questions_per_page
        if prev_start < 0:
            prev_start = 0
        self.current_question_index = prev_start
        return True

    def save_report(self):
        """Сохранение отчета в БД и экспорт в Excel"""
        try:
            # Очищаем имя от запрещённых символов
            clean_form_name = sanitize_filename(self.current_report_data['form_name'])
            clean_month = sanitize_filename(self.current_report_data['month'])
            clean_year = str(self.current_report_data['year'])

            report_name = f"{clean_form_name}_{clean_month}_{clean_year}"

            file_path = create_excel_report(
                report_name=report_name,
                form_name=self.current_report_data['form_name'],
                month=self.current_report_data['month'],
                year=self.current_report_data['year'],
                answers=self.answers_list
            )

            report_id = save_report_to_db(self.current_report_data, self.answers_list, file_path)

            return True, os.path.basename(file_path)

        except Exception as e:
            return False, str(e)

    def get_all_reports_from_db(self):
        """Получить все отчеты из БД"""
        return get_all_reports()

    def get_report_from_db(self, report_id):
        """Получить конкретный отчет из БД"""
        return get_report_by_id(report_id)

    def delete_report_from_db(self, report_id):
        """Удалить отчет из БД"""
        try:
            delete_report(report_id)
            return True, "Отчет удален"
        except Exception as e:
            return False, str(e)

    def export_report_to_word(self, report_data):
        """Экспортировать отчет в Excel заново"""
        try:
            # Очищаем имя от запрещённых символов
            clean_form_name = sanitize_filename(report_data['form_name'])
            clean_month = sanitize_filename(report_data['month'])
            clean_year = str(report_data['year'])

            report_name = f"{clean_form_name}_{clean_month}_{clean_year}"

            file_path = create_excel_report(
                report_name=report_name,
                form_name=report_data['form_name'],
                month=report_data['month'],
                year=report_data['year'],
                answers=report_data['answers']
            )

            return True, os.path.basename(file_path)
        except Exception as e:
            return False, str(e)