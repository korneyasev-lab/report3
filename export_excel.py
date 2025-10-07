# -*- coding: utf-8 -*-
"""
export_excel.py — создаёт Excel отчёты на основе шаблонов
"""

import os
from datetime import datetime
import openpyxl


def _get_templates_dir():
    templates = os.path.join(os.getcwd(), "шаблоны")
    os.makedirs(templates, exist_ok=True)
    return templates


def _reports_dir():
    reports = os.path.join(os.getcwd(), "отчеты")
    os.makedirs(reports, exist_ok=True)
    return reports


def sanitize_filename(name):
    bad = r'\/:*?"<>|%#'
    return "".join(("_" if c in bad else c) for c in name).strip() or "report"


def hide_empty_comment_rows(ws):
    """Скрывает все пустые строки (кроме разделителей)"""
    # Находим строку с датой
    date_row = None
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row_idx, 1).value
        if cell_value and "Дата создания отчета:" in str(cell_value):
            date_row = row_idx
            break

    # Скрываем все пустые строки
    for row_idx in range(3, ws.max_row + 1):
        # Пропускаем строку перед датой (разделитель)
        if date_row and row_idx == date_row - 1:
            continue

        # Проверяем пустая ли строка
        cell_a = ws.cell(row_idx, 1).value
        cell_b = ws.cell(row_idx, 2).value

        if (not cell_a or not str(cell_a).strip()) and (not cell_b or not str(cell_b).strip()):
            ws.row_dimensions[row_idx].hidden = True


def create_excel_report(report_name, form_name, month, year, answers):
    """Создаёт отчёт из шаблона"""

    # Загружаем шаблон
    template_path = os.path.join(_get_templates_dir(), "отчет.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Заполняем шапку
    ws['A1'] = report_name

    # Заполняем вопросы в готовые строки
    for i, item in enumerate(answers or []):
        # Данные
        if isinstance(item, dict):
            q = str(item.get("question_text", ""))
            a = str(item.get("answer_yes_no", ""))
            c = str(item.get("comment", ""))
        else:
            q = str(getattr(item, "question_text", ""))
            a = str(getattr(item, "answer_yes_no", ""))
            c = str(getattr(item, "comment", ""))

        # Строки для этого вопроса
        question_row = 3 + i * 2
        comment_row = 4 + i * 2

        # Заполняем
        ws.cell(question_row, 1).value = q.strip()
        ws.cell(question_row, 2).value = a.strip()
        ws.cell(comment_row, 1).value = c.strip() if c.strip() else ""

    # Скрываем пустые строки комментариев
    hide_empty_comment_rows(ws)

    # Заполняем дату
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row_idx, 1).value
        if cell_value and "Дата создания отчета:" in str(cell_value):
            ws.cell(row_idx, 1).value = f"Дата создания отчета: {datetime.now().strftime('%d.%m.%Y')}"
            break

    # Сохраняем
    reports_dir = _reports_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = sanitize_filename(f"{report_name}_{timestamp}.xlsx")
    out_path = os.path.join(reports_dir, safe)

    wb.save(out_path)

    # Показываем путь
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Отчёт сохранён", f"Файл сохранён в:\n\n{out_path}")
        root.destroy()
    except:
        pass

    return out_path